'''This module perform several function with the aim of processing and bulk uploading products
from an xlsx file.
The first function reads the data from the xlsx file. The second function checks for typo or syntax error in your
file that could result in unwanted uploads or uploads with wrong information.
The third file uploads your products to your Woocommerce store through API calls.

Written by Simone Onorato
Kingston,Ontario
support: simon.onorato@queensu.ca'''

from woocommerce import API
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from constants import constantData
from scp_com import uploadTempFiles as uploadImm
from scp_com import is_url_image as imgCheck

#TESTING
file_location= '.\\templates\\input_template.xlsx'


#This function reads data from the xsls file and imports it into memory
def readData(inputFile):
    print('-----------------------------------')
    print('\nReading data from your file... \n')
    #Global function variables
    keyRow = 2
    valueRow = 3
    columnRange = 29

    #This list will contain the whole xsls in form of dictionaries
    completeData=[]

    # opening input workbook
    inputWb = load_workbook(inputFile)
    inputWs = inputWb.active

    #reading data
    currentValueRow=valueRow
    while inputWs[f'A{currentValueRow}'].value is not None:
        productData={}
        currentColumn=1

        for cell in range(columnRange): #range has to be total number of columns assigned to data entry
            #Detecting Data Coordinates and reading data from the cells
            currentLetter = get_column_letter(currentColumn) #detecting the current column
            currentKey = inputWs[f'{currentLetter}{keyRow}'].value  # taking the key from a cell
            currentKey = str(currentKey)  # converting cell key to a string
            currentValue = inputWs[f'{currentLetter}{currentValueRow}'].value #taking the value from a cell
            currentValue = str(currentValue) #converting cell value to a string
            if currentValue == 'None': #sanitizing inputs
                currentValue = ''
            productData[currentKey] = currentValue #Appending data to a dictionary
            currentColumn += 1 #Increasing the column count by 1

        #appending a product in form of a dictionary to the complete list of products
        completeData.append(productData)
        currentValueRow +=1 #increasing value row count by 1
    print(f'File imported correctly. {currentValueRow-2} products imported. \n')
    return completeData #returning a list of dictionaries with the whole xlsx data


#This function checks the data imported in memory for several typos and errors pointing out where the errors are to
#be found in the xlsx file
def checkFile(completeData):
    print('-----------------------------------')
    print('Checking your file for errors \n')

    #Defining global variables
    woocommerce_keys_to_check = ["title", "tags", "ean", "sku", "shop-category", "quantity", "price", "sale-price",
                                 "shipping", "all-images", "var1", "var2"]
    woocommerce_categories_to_check = constantData.shopCategories
    woocommerce_brands_to_check = constantData.shopBrands

    #creating an empty error list
    errorList=[]

    #iterarting across dictionaries
    for dictionary in completeData:
        #checking for empty cell errors
        for key in dictionary:
            if key in woocommerce_keys_to_check:
                value=dictionary.get(key, default)
                if value == '':
                    errorList.append(f'Value for key {key} is empty on row {dictionary}.')

            #checking for shop category errors
            if key == 'shop-category':
                value=dictionary.get(key, default)
                if value not in woocommerce_categories_to_check.values():
                    errorList.append(f'The shop category {value} for product on row {dictionary} '
                                     f'is not written in the correct way. Check that the category name you have written'
                                     f'matches the category name on the website precisely.')

            #checking for shop brands errors
            if key == 'brand':
                value=dictionary.get(key, default)
                if value not in woocommerce_brands_to_check.values():
                    errorList.append(f'The shop brand {value} for product on row {dictionary} '
                                     f'is not written in the correct way. Check that the category name you have written'
                                     f'matches the category name on the website precisely.')

            #checking for float number errors
            if key == "price" or key == "sale price":
                value = dictionary.get(key, default)
                if value != '':
                    if value.find('.') == -1:
                        errorList.append(f'The variables for {key} on line {dictionary} do not contain a "."'
                                         f' between numbers. It is common to miss this error if you inadvertently used ",".')

            #cheking for variable errors
            if key == "var1" or key == "var2":
                value=dictionary.get(key, default)
                if value != '':
                    if value.find(':') == -1:
                        errorList.append(f'The variables for {key} on line {dictionary} do not contain a ":".'
                                         f'It is common to miss this error if you inadvertently used ";".')
    #Creating a report if errors were found or printing a 'no errors detected' statement.
    if len(errorList) != 0:
        errorNumber=0
        for error in errorList:
            print('Some errors were found, here is a list:')
            print('-------------------------------------------')
            print (f'{errorNumber+1}. {error}')
            print('\nDo you still want to upload your products?\n')
            uploadAnyway=str(input('type y or n: '))
            if uploadAnyway == 'n' or uploadAnyway == 'N':
                return ('Error detected')
    else:
        print('Congratulation! No errors were found.')

#upload products
def uploadProducts(completeData):
    print('-----------------------------------')
    print('\n Starting product upload. \n')
    productsToUpload=len(completeData)
    print(f'Total products to upload: {productsToUpload}\n')

    #Authenticate
    print('Establishing connection and authenticating with your marketplace. \n')
    '''
    wcapi = API(
        url=constantData.shop_url,
        consumer_key=constantData.consumer_key,
        consumer_secret=constantData.consumer_secret,
        version=constantData.version
    )
    '''

    #product upload
    for product in completeData:
        print(product)
        currentProduct=1
        currentTitle= product.get('TITLE', 'default')
        print(f'Uploading product {currentProduct} of {productsToUpload}. \n'
          f'Product Title: {currentTitle}. \n')

        #preparing the data
        productQuantity=int(product.get('QUANTITY',''))
        product['QUANTITY'] = productQuantity

        #Checking if simple or variable listing
        var1= product.get('VARIABLE_1','')
        if var1 == '':
            listing_type="simple"
        else:
            listing_type="variable"

        #creating categories and brands list
        categories_list=product.get('SHOP_CATEGORY','').split(',')
        print('Category list:', categories_list)
        brand_list=product.get('BRAND','').split(',')
        print('brand list:', brand_list)
        if categories_list == '':
            print('You forgot to add categories to the listing title')
            print(f'Listing Title: {currentTitle}')
            print('Upload will be done with no categories')
        if brand_list == '':
            print('You forgot to add brands to the listing title')
            print(f'Listing Title: {currentTitle}')
            print('Upload will be done with no brands')

        #Converting categories and brands to their ID accepted by the website
        categories_brands = []
        if categories_list != '':
            for key in categories_list:
                if key in constantData.shopCategories:
                    category_id= constantData.shopCategories.get(key,'')
                    id={"id":category_id}
                    categories_brands.append(id)
        if brand_list != '':
            for key in brand_list:
                if key in constantData.shopCategories:
                    category_id= constantData.shopCategories.get(key,'')
                    id={"id":category_id}
                    categories_brands.append(id)

        #Uploading images to the server
        print('Upload your listing images to the server\n')
        uploadImm()



        #Check if images are online otherwise upload them on the server
        images_list = product.get('IMAGES','').split(',')
        total_imgs=len(images_list)
        imgs_exist=False
        img_exist_counter=0
        for image in images_list:
            img_url_jpg=constantData.remote_image_link+image+'.jpg'
            img_url_png = constantData.remote_image_link + image + '.png'
            exist1=imgCheck(img_url_jpg)
            exist2 = imgCheck(img_url_png)
            if exist1 or exist2 is True:
                img_exist_counter += 1
        if img_exist_counter == total_imgs:
            imgs_exist = True

        if imgs_exist is False:
            print('It seems like you forgot to upload the images onto the server..')
            print('He he he... would you like to upload them now?')
            print()
            userInput=str(input('y=yes n=no: '))
            if userInput == 'y' or userInput == 'Y':
                uploadImm()
            else:
                print('Sorry, can not upload a listing if the images are not on the server =(')
                return False

        # create images payload
        payload_images = []
        for image in images_list:
            img_url = constantData.remote_image_directory + image + '.jpg'
            img_src = {"src": img_url}
            payload_images.append(img_src)

        # create variation payload
        # var1 - analyzing
        var1 = product.get('VARIATION 1','')
        if var1 != '':
            var1_name_pos = var1.find(':')
            var1_name_plus = var1[0:var1_name_pos + 1]
            var1_clean = var1.replace(var1_name_plus, '')

            var1_name = var1[0:var1_name_pos]
            var1_list = var1_clean.split(',')
        # var2 - analyzing
        var2 = product.get('VARIATION 2','')
        if var2 != '':
            var2_name_pos = var2.find(':')
            var2_name_plus = var2[0:var2_name_pos + 1]
            var2_clean = var2.replace(var2_name_plus, '')

            var2_name = var2[0:var2_name_pos]
            var2_list = var2_clean.split(',')

        # detect status
        listingStatus=product.get('STATE','')
        if listingStatus != 'draft' or listingStatus != 'DRAFT':
            status = "publish"
        else:
            status = "draft"

        #Defining other fields
        productSku=product.get('SKU','')
        productPrice= product.get('PRICE','')
        productSalePrice= product.get('SALE PRICE','')

        #Generating descriptions
        #Generating short description
        shortDescription=''
        descriptionTitle=product.get("SHORT TITLE","")
        if descriptionTitle != '':
            sDescription = (f'{product.get("SHORT TITLE","")}{product.get("BULLET POINT 1","")}{product.get("BULLET POINT 2","")}'
                            f'{product.get("BULLET POINT 3","")}{product.get("BULLET POINT 4","")}{product.get("BULLET POINT 5","")}'
                            f'{product.get("DIMENSIONS","")}{product.get("NET WEIGHT","")}{product.get("CAPACITY","")}')
            print('short description generated')

        space = ''
        longDescription=''
        if product.get("LONG TITLE","") != '' or product.get("TEXT1","") != '':
            space = '\n<br>\n<br>\n'
            longDescription = (f'{product.get("LONG TITLE","")}{product.get("TEXT1","")}{product.get("PARAGRAPH TITLE","")}'
                        f'{product.get("TEXT2","")}{space}{shortDescription}')
            print('long description generated')

        # uploading a product

        payload_listing = {"name": currentTitle, "type": listing_type, "status": status,
                           "description": longDescription,
                           "short_description": shortDescription, "sku": productSku,
                           "regular_price": productPrice,
                           "sale_price": productSalePrice, "manage_stock": True,
                           "stock_quantity": productQuantity, "categories": categories_brands,
                           "images": payload_images,
                           "attributes": [{"name": var1_name, "variation": True, "visible": True,
                                           "options": var1_list},
                                          {"name": var2_name, "variation": True, "visible": True,
                                           "options": var2_list}]}

        req = wcapi.post("products", payload_listing)

        # print('product upload', req.status_code)
        # print(req.text)
        req = req.json()
        print(req)

        if 'code' in req.keys():
            # Check if Sku Already Used
            if req["code"] == 'product_invalid_sku':
                print()
                print()
                print('For Product', rowData["title"])
                print('SKU already used, it will not be uploaded')


            else:
                # Create Variations, this command generate all possible variations without assigning pictures to specific variations
                listing_id = req["id"]
                var_sku = product.get('SKU','') + "_var"
                variations = {"regular_price": product.get('PRICE',''), "sku": var_sku}
                req = wcapi.post("products/" + str(listing_id) + "/variations", variations)
                # print('var upload', req.status_code)
                # print(req.text)
                req = req.json()

        print('Upload compleated.')
        print()
        print()



